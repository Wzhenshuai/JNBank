import json
import time

__author__ = 'admin'
import openpyxl
from openpyxl.styles import Color, Font, Alignment, Border, Side
import openpyxl.styles as sty

border_ww = Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
font_ww = Font(name='仿宋', size=14, bold=True)
font_content_ww = Font(name='仿宋', size=13, bold=False)
alignment_ww = Alignment(horizontal='center', vertical='center')
title_fill = sty.PatternFill(fill_type="solid", fgColor="A9A9A9")

def create_table_sheet(wb, title):
    sheet2 = wb.create_sheet(str(0))
    sheet2.title = title
    sheet2.cell(row=1, column=1, value=str("中文名称")).fill=title_fill
    sheet2.cell(row=1, column=2, value=str("1"))
    sheet2.cell(row=1, column=3, value=str("2"))
    sheet2.cell(row=1, column=4, value=str("3"))
    sheet2.cell(row=1, column=5, value=str("4"))
    sheet2.cell(row=1, column=6, value=str("5"))
    sheet2.cell(row=1, column=7, value=str("6"))
    sheet2.cell(row=1, column=8, value=str("7"))
    sheet2.cell(row=1, column=9, value=str("8"))
    sheet2.cell(row=1, column=10, value=str("9"))

    sheet2.cell(row=2, column=1, value=str("英文名称")).fill=title_fill
    sheet2.cell(row=2, column=2, value=str("1"))
    sheet2.cell(row=2, column=3, value=str("2"))
    sheet2.cell(row=2, column=4, value=str("3"))
    sheet2.cell(row=2, column=5, value=str("4"))
    sheet2.cell(row=2, column=6, value=str("5"))
    sheet2.cell(row=2, column=7, value=str("6"))
    sheet2.cell(row=2, column=8, value=str("7"))
    sheet2.cell(row=2, column=9, value=str("8"))
    sheet2.cell(row=2, column=10, value=str("9"))

    sheet2.cell(row=3, column=1, value=str("说明")).fill=title_fill
    sheet2.cell(row=3, column=2, value=str("1"))
    sheet2.cell(row=3, column=3, value=str("2"))
    sheet2.cell(row=3, column=4, value=str("3"))
    sheet2.cell(row=3, column=5, value=str("4"))
    sheet2.cell(row=3, column=6, value=str("5"))
    sheet2.cell(row=3, column=7, value=str("6"))
    sheet2.cell(row=3, column=8, value=str("7"))
    sheet2.cell(row=3, column=9, value=str("8"))
    sheet2.cell(row=3, column=10, value=str("9"))

    sheet2.merge_cells('B1:J1')
    sheet2.merge_cells('B2:J2')
    sheet2.merge_cells('B3:J3')

    sheet2.cell(row=4, column=1, value=str("序号")).fill=title_fill
    sheet2.cell(row=4, column=2, value=str("字段名称")).fill=title_fill
    sheet2.cell(row=4, column=3, value=str("字段代码")).fill=title_fill
    sheet2.cell(row=4, column=4, value=str("字段类型")).fill=title_fill
    sheet2.cell(row=4, column=5, value=str("字段长度")).fill=title_fill
    sheet2.cell(row=4, column=6, value=str("精确度")).fill=title_fill
    sheet2.cell(row=4, column=7, value=str("主键标识")).fill=title_fill
    sheet2.cell(row=4, column=8, value=str("是否为字典")).fill=title_fill
    sheet2.cell(row=4, column=9, value=str("字典内容")).fill=title_fill
    sheet2.cell(row=4, column=10, value=str("备注")).fill=title_fill

    sheet2.row_dimensions[1].height = 20.0
    sheet2.row_dimensions[2].height = 20.0
    sheet2.row_dimensions[3].height = 20.0

    title_cells_2 = ['A','B','C','D','E','F','G','H','I','J']
    for i in range(1,4):
        sheet2['A'+str(i)].border = border_ww
        sheet2['A'+str(i)].font = font_ww
        sheet2['A'+str(i)].alignment = alignment_ww
        for j in range(2,11):
            sheet2.cell(i,j).border = border_ww
            sheet2.cell(i,j).font = font_content_ww
            sheet2.cell(i,j).alignment = alignment_ww
    for i in title_cells_2:
        sheet2[i+'4'].border = border_ww
        sheet2[i+'4'].font = font_ww
        sheet2[i+'4'].alignment = alignment_ww
        sheet2.column_dimensions[i].width = 15.0
    return sheet2


def create_xlsx(sys_name, data_file_path, target_dir_path):
    target_file_name = sys_name + "表结构-技术-"+time.strftime('%Y%m%d%H%M%S',time.localtime(time.time())) + ".xlsx"
    file = open(data_file_path, "rb")
    file_content = file.read()
    file_json = json.loads(file_content.decode("utf-8"))

    wb = openpyxl.Workbook()
    sheet0 = wb.active
    sheet0.title = '变更记录'
    sheet0.cell(row=1, column=1, value=str("修改日期")).fill=title_fill
    sheet0.cell(row=1, column=2, value=str("表名称")).fill=title_fill
    sheet0.cell(row=1, column=3, value=str("修改内容")).fill=title_fill
    sheet0.cell(row=1, column=4, value=str("修改说明")).fill=title_fill
    sheet0.cell(row=1, column=5, value=str("修改人")).fill=title_fill
    sheet0.cell(row=1, column=6, value=str("需求类型")).fill=title_fill
    sheet0.cell(row=1, column=7, value=str("备注")).fill=title_fill

    sheet1 = wb.create_sheet(str(0))
    sheet1.title = '索引'
    sheet1.cell(row=1, column=1, value=str("系统名称")).fill=title_fill
    sheet1.cell(row=1, column=2, value=str("中文名称")).fill=title_fill
    sheet1.cell(row=1, column=3, value=str("英文名称")).fill=title_fill
    sheet1.cell(row=1, column=4, value=str("供数方式")).fill=title_fill
    sheet1.cell(row=1, column=5, value=str("供数原因")).fill=title_fill
    sheet1.cell(row=1, column=6, value=str("下档文件名称")).fill=title_fill
    sheet1.cell(row=1, column=7, value=str("法人与村镇标识")).fill=title_fill
    sheet1.cell(row=1, column=8, value=str("负责人")).fill=title_fill
    sheet1.cell(row=1, column=9, value=str("备注")).fill=title_fill

    sheet0.row_dimensions[1].height = 20.0
    sheet1.row_dimensions[1].height = 20.0

    title_cells_0 = ['A','B','C','D','E','F','G']
    for i in title_cells_0:
        sheet0[i+'1'].border = border_ww
        sheet0[i+'1'].font = font_ww
        sheet0[i+'1'].alignment = alignment_ww
        sheet0.column_dimensions[i].width = 15.0
    title_cells_1 = ['A','B','C','D','E','F','G','H','I']
    for i in title_cells_1:
        sheet1[i+'1'].border = border_ww
        sheet1[i+'1'].font = font_ww
        sheet1[i+'1'].alignment = alignment_ww
        sheet1.column_dimensions[i].width = 15.0

    table_no = 1
    font_hyperlink_style = Font(underline='single', color='0066cc')
    for table in file_json:
        table_no += 1
        table_name_zh = table["table_name_zh"]
        table_name_cn = table["table_name_cn"]
        sheet1.cell(table_no, 1, sys_name)
        sheet1.cell(table_no, 2, table_name_zh)
        sheet1.cell(table_no, 3, table_name_cn).hyperlink = target_file_name + "#" + str(table_name_cn)+"!A1"
        sheet1["C" + str(table_no)].font = font_hyperlink_style

        table_sheet = create_table_sheet(wb, table_name_cn)
        table_sheet.cell(1,2).value = table_name_zh
        table_sheet.cell(2,2).value = table_name_cn
        table_sheet.cell(3,2).value = ""

        column_datas = table["column_datas"]
        row_no = 4
        for column in column_datas:
            row_no += 1
            table_sheet.cell(row_no, 1, row_no - 4)
            if column.__contains__("column_name_zh"):
                table_sheet.cell(row_no, 2, column["column_name_zh"])
            if column.__contains__("column_name_cn"):
                table_sheet.cell(row_no, 3, column["column_name_cn"])
            if column.__contains__("column_type"):
                table_sheet.cell(row_no, 4, column["column_type"])
            if column.__contains__("column_length"):
                table_sheet.cell(row_no, 5, column["column_length"])
            if column.__contains__("column_precision"):
                table_sheet.cell(row_no, 6, column["column_precision"])
            if column.__contains__("column_is_pk"):
                table_sheet.cell(row_no, 7, column["column_is_pk"])
            if column.__contains__("column_is_dict"):
                table_sheet.cell(row_no, 8, column["column_is_dict"])
            if column.__contains__("column_dict_content"):
                table_sheet.cell(row_no, 9, column["column_dict_content"])
            if column.__contains__("column_note"):
                table_sheet.cell(row_no, 10, column["column_note"])
    wb.save(target_dir_path + target_file_name)
    print(target_dir_path + target_file_name + "写入数据成功！")

if __name__ == '__main__':
    # TODO(王廷国):填写具体的系统名称
    sys_name = "财务系统"
    # TODO(王廷国):填写系统对应表结构json数据文件路径
    data_file_path = "D:/开发快捷/财务系统.txt"
    # TODO(王廷国):填写输出结果文件存放路径
    target_dir_path = "D:/开发快捷/"

    # 创建目标xlsx文件
    create_xlsx(sys_name, data_file_path, target_dir_path)
